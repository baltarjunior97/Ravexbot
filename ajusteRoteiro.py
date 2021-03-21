from openpyxl import load_workbook
import openpyxl
from openpyxl.worksheet import worksheet
from datetime import date

hoje = date.today()

d1 = hoje.strftime("%d.%m.%Y")

wb = load_workbook('C:\\Users\djalm\Downloads\RODOMAIKE.xlsx')

sheet = wb['Plan1'] 

for i in range(0, 300):
    x = sheet.cell(row=i+1, column=13).value
    if x == None:
        break
    else:
        sheet.cell(row=i+1, column=13).value = None

worksheet.Worksheet.delete_cols(sheet, 3)
worksheet.Worksheet.delete_cols(sheet, 4)
worksheet.Worksheet.delete_cols(sheet, 4)
worksheet.Worksheet.delete_cols(sheet, 7)
worksheet.Worksheet.delete_cols(sheet, 7)


roteiro =str("C://Users/djalm/Downloads/roteiro{}{}".format(d1, ".xlsx"))    
openpyxl.Workbook.save(wb, filename=roteiro)