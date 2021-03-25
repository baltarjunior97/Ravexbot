from openpyxl import load_workbook
import openpyxl
from openpyxl.worksheet import worksheet
from datetime import date
from openpyxl.styles import PatternFill

# Variaveis
carro = []
com = []
fim = []
notas1 = []
notas2 = []
notas3 = []
notas4 = []
notas5 = []
notas6 = []
feitas = []
hoje = date.today().strftime("%d.%m.%Y")
roteiro = str("C://Users/user/Local/NomeFinaldoArquivo")


# funções

def semCarimbo():
    nNota1 = 0
    car = str(input('Qual carro? '))
    while True:
        if car not in carro:
            print('Carros Disponiveis:')
            for y in range(0, len(carro)):
                print(carro[y])
            car = str(input('Qual carro? '))
        elif car in carro:
            break
    xpos = carro.index(car)
    if xpos == 0:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas1:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas1:
                break
    if xpos == 1:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas2:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas2:
                break
    if xpos == 2:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas3:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas3:
                break
    if xpos == 3:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas4:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas4:
                break
    if xpos == 4:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas5:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas5:
                break
    if xpos == 5:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas6:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas6:
                break
    note = 'sem carimbo'

    if xpos == 0:
        for i in range(com[xpos], fim[xpos]):
            if ws1.cell(row=i, column=3).value == nNota1:
                ws1.cell(row=i, column=8).value = note
    if xpos == 1:
        for i in range(com[xpos], fim[xpos]):
            if ws2.cell(row=i, column=3).value == nNota1:
                ws2.cell(row=i, column=8).value = note
    if xpos == 2:
        for i in range(com[xpos], fim[xpos]):
            if ws3.cell(row=i, column=3).value == nNota1:
                ws3.cell(row=i, column=8).value = note
    if xpos == 3:
        for i in range(com[xpos], fim[xpos]):
            if ws4.cell(row=i, column=3).value == nNota1:
                ws4.cell(row=i, column=8).value = note
    if xpos == 4:
        for i in range(com[xpos], fim[xpos]):
            if ws5.cell(row=i, column=3).value == nNota1:
                ws5.cell(row=i, column=8).value = note
    if xpos == 5:
        for i in range(com[xpos], fim[xpos]):
            if ws6.cell(row=i, column=3).value == nNota1:
                ws6.cell(row=i, column=8).value = note


def reentrega():
    nNota1 = 0
    car = str(input('Qual carro teve a reentrega? '))
    while True:
        if car not in carro:
            print('Carros Disponiveis:')
            for y in range(0, len(carro)):
                print(carro[y])
            car = str(input('Qual carro teve a reentrega? '))
        elif car in carro:
            break
    xpos = carro.index(car)
    if xpos == 0:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas1:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas1:
                break
    if xpos == 1:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas2:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas2:
                break
    if xpos == 2:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas3:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas3:
                break
    if xpos == 3:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas4:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas4:
                break
    if xpos == 4:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas5:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas5:
                break
    if xpos == 5:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas6:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas6:
                break
    note = str(input('Digite o motivo: '))

    if xpos == 0:
        for i in range(com[xpos], fim[xpos]):
            if ws1.cell(row=i, column=3).value == nNota1:
                ws1.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws1.cell(row=i, column=y).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
    if xpos == 1:
        for i in range(com[xpos], fim[xpos]):
            if ws2.cell(row=i, column=3).value == nNota1:
                ws2.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws2.cell(row=i, column=y).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
    if xpos == 2:
        for i in range(com[xpos], fim[xpos]):
            if ws3.cell(row=i, column=3).value == nNota1:
                ws3.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws3.cell(row=i, column=y).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
    if xpos == 3:
        for i in range(com[xpos], fim[xpos]):
            if ws4.cell(row=i, column=3).value == nNota1:
                ws4.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws4.cell(row=i, column=y).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
    if xpos == 4:
        for i in range(com[xpos], fim[xpos]):
            if ws5.cell(row=i, column=3).value == nNota1:
                ws5.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws5.cell(row=i, column=y).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
    if xpos == 5:
        for i in range(com[xpos], fim[xpos]):
            if ws6.cell(row=i, column=3).value == nNota1:
                ws6.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws6.cell(row=i, column=y).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")


def devolucao():
    nNota1 = 0
    dev = 0

    dev = str(input('Devolução parcial(1) ou devolução total(2)? '))
    while True:
        if dev == '1' or dev == 'devolução parcial' or dev == 'devolucao parcial':
            dev = 'devolução parcial'
            break
        elif dev == '2' or dev == 'devolução total' or dev == 'devoucao total':
            dev = 'devolução total'
            break
        else:
            dev = str(input('Devolução parcial(1) ou devolução total(2)? '))

    car = str(input(f'Qual carro teve a {dev}? '))
    while True:
        if car not in carro:
            print('Carros Disponiveis:')
            for y in range(0, len(carro)):
                print(carro[y])
            car = str(input(f'Qual carro teve a {dev}? '))
        elif car in carro:
            break
    xpos = carro.index(car)
    if xpos == 0:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas1:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas1:
                break
    if xpos == 1:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas2:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas2:
                break
    if xpos == 2:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas3:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas3:
                break
    if xpos == 3:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas4:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas4:
                break
    if xpos == 4:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas5:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas5:
                break
    if xpos == 5:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas6:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas6:
                break
    note = str(input('Digite o motivo: '))
    if xpos == 0:
        for i in range(com[xpos], fim[xpos]):
            if ws1.cell(row=i, column=3).value == nNota1:
                ws1.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws1.cell(row=i, column=y).fill = PatternFill(fgColor="00FF0000", fill_type="solid")
    if xpos == 1:
        for i in range(com[xpos], fim[xpos]):
            if ws2.cell(row=i, column=3).value == nNota1:
                ws2.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws2.cell(row=i, column=y).fill = PatternFill(fgColor="00FF0000", fill_type="solid")
    if xpos == 2:
        for i in range(com[xpos], fim[xpos]):
            if ws3.cell(row=i, column=3).value == nNota1:
                ws3.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws3.cell(row=i, column=y).fill = PatternFill(fgColor="00FF0000", fill_type="solid")
    if xpos == 3:
        for i in range(com[xpos], fim[xpos]):
            if ws4.cell(row=i, column=3).value == nNota1:
                ws4.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws4.cell(row=i, column=y).fill = PatternFill(fgColor="00FF0000", fill_type="solid")
    if xpos == 4:
        for i in range(com[xpos], fim[xpos]):
            if ws5.cell(row=i, column=3).value == nNota1:
                ws5.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws5.cell(row=i, column=y).fill = PatternFill(fgColor="00FF0000", fill_type="solid")
    if xpos == 5:
        for i in range(com[xpos], fim[xpos]):
            if ws6.cell(row=i, column=3).value == nNota1:
                ws6.cell(row=i, column=8).value = note
                for y in range(1, 9):
                    ws6.cell(row=i, column=y).fill = PatternFill(fgColor="00FF0000", fill_type="solid")


def anotacao():
    nNota1 = 0
    car = str(input('Qual carro vc quer fazer a anotaçao? '))
    while True:
        if car not in carro:
            print('Carros Disponiveis:')
            for y in range(0, len(carro)):
                print(carro[y])
            car = str(input('Qual carro vc quer fazer a anotaçao? '))
        elif car in carro:
            break
    xpos = carro.index(car)
    if xpos == 0:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas1:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas1:
                break
    if xpos == 1:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas2:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas2:
                break
    if xpos == 2:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas3:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas3:
                break
    if xpos == 3:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas4:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas4:
                break
    if xpos == 4:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas5:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas5:
                break
    if xpos == 5:
        nNota1 = int(input(f'Qual nota do carro {car}?'))
        while True:
            if nNota1 not in notas6:
                nNota1 = int(input(f'Qual nota do carro {car}?'))
            elif nNota1 in notas6:
                break
    note = str(input('Digite a anotaçao: '))

    if xpos == 0:
        for i in range(com[xpos], fim[xpos]):
            if ws1.cell(row=i, column=3).value == nNota1:
                ws1.cell(row=i, column=8).value = note
    if xpos == 1:
        for i in range(com[xpos], fim[xpos]):
            if ws2.cell(row=i, column=3).value == nNota1:
                ws2.cell(row=i, column=8).value = note
    if xpos == 2:
        for i in range(com[xpos], fim[xpos]):
            if ws3.cell(row=i, column=3).value == nNota1:
                ws3.cell(row=i, column=8).value = note
    if xpos == 3:
        for i in range(com[xpos], fim[xpos]):
            if ws4.cell(row=i, column=3).value == nNota1:
                ws4.cell(row=i, column=8).value = note
    if xpos == 4:
        for i in range(com[xpos], fim[xpos]):
            if ws5.cell(row=i, column=3).value == nNota1:
                ws5.cell(row=i, column=8).value = note
    if xpos == 5:
        for i in range(com[xpos], fim[xpos]):
            if ws6.cell(row=i, column=3).value == nNota1:
                ws6.cell(row=i, column=8).value = note


def check():
    for i in range(0, len(carro)):
        xpos = i
        while True:
            try:
                x = int(input(f'até que nota o {carro[i]} fez? '))
                if x >= com[i] and x <= fim[i]:
                    feitas.append(x)
                    break
            except Exception:
                continue
        if xpos == 0:
            for i in range(com[xpos], feitas[xpos]):
                for y in range(1, 9):
                    if ws1.cell(row=i, column=y).fill.start_color.index == '00000000':
                        ws1.cell(row=i, column=y).fill = PatternFill(fgColor="0000FF00", fill_type="solid")
        if xpos == 1:
            for i in range(com[xpos], feitas[xpos]):
                for y in range(1, 9):
                    if ws2.cell(row=i, column=y).fill.start_color.index == '00000000':
                        ws2.cell(row=i, column=y).fill = PatternFill(fgColor="0000FF00", fill_type="solid")
        if xpos == 2:
            for i in range(com[xpos], feitas[xpos]):
                for y in range(1, 9):
                    if ws3.cell(row=i, column=y).fill.start_color.index == '00000000':
                        ws3.cell(row=i, column=y).fill = PatternFill(fgColor="0000FF00", fill_type="solid")
        if xpos == 3:
            for i in range(com[xpos], feitas[xpos]):
                for y in range(1, 9):
                    if ws4.cell(row=i, column=y).fill.start_color.index == '00000000':
                        ws4.cell(row=i, column=y).fill = PatternFill(fgColor="0000FF00", fill_type="solid")
        if xpos == 4:
            for i in range(com[xpos], feitas[xpos]):
                for y in range(1, 9):
                    if ws5.cell(row=i, column=y).fill.start_color.index == '00000000':
                        ws5.cell(row=i, column=y).fill = PatternFill(fgColor="0000FF00", fill_type="solid")
        if xpos == 5:
            for i in range(com[xpos], feitas[xpos]):
                for y in range(1, 9):
                    if ws6.cell(row=i, column=y).fill.start_color.index == '00000000':
                        ws6.cell(row=i, column=y).fill = PatternFill(fgColor="0000FF00", fill_type="solid")


# Formatação final da planilha principal
def formatPlan():
    lin = 0
    col = 0

    for i in range(com[0], fim[0]):
        lin = lin + 1
        col = 0
        for j in range(1, 10):
            col = col + 1
            plan = ws1.cell(row=i, column=j)
            aba.cell(row=lin + 1, column=col).value = plan.value
            if aba.cell(row=lin + 1, column=col).fill.start_color.index == '00000000':
                cor = ws1.cell(row=i, column=j).fill.start_color.index
                if cor != '00000000':
                    aba.cell(row=lin + 1, column=col).fill = PatternFill(fgColor=cor, fill_type='solid')

    col = 0
    for j in range(1, 9):
        col = col + 1
        aba.cell(row=lin + 2, column=col).fill = PatternFill(fgColor="808080", fill_type="solid")
    try:
        for i in range(com[1], fim[1]):
            lin = lin + 1
            col = 0
            for j in range(1, 10):
                col = col + 1
                plan = ws2.cell(row=i, column=j)
                aba.cell(row=lin + 2, column=col).value = plan.value
                if aba.cell(row=lin + 2, column=col).fill.start_color.index == '00000000':
                    cor = ws2.cell(row=i, column=j).fill.start_color.index
                    if cor != '00000000':
                        aba.cell(row=lin + 2, column=col).fill = PatternFill(fgColor=cor, fill_type='solid')

        col = 0
        for j in range(1, 9):
            col = col + 1
            aba.cell(row=lin + 3, column=col).fill = PatternFill(fgColor="808080", fill_type="solid")
    except Exception:
        pass

    try:
        for i in range(com[2], fim[2]):
            lin = lin + 1
            col = 0
            for j in range(1, 10):
                col = col + 1
                plan = ws3.cell(row=i, column=j)
                aba.cell(row=lin + 3, column=col).value = plan.value
                if aba.cell(row=lin + 3, column=col).fill.start_color.index == '00000000':
                    cor = ws3.cell(row=i, column=j).fill.start_color.index
                    if cor != '00000000':
                        aba.cell(row=lin + 3, column=col).fill = PatternFill(fgColor=cor, fill_type='solid')

        col = 0
        for j in range(1, 9):
            col = col + 1
            aba.cell(row=lin + 4, column=col).fill = PatternFill(fgColor="808080", fill_type="solid")
    except Exception:
        pass
    try:
        for i in range(com[3], fim[3]):
            lin = lin + 1
            col = 0
            for j in range(1, 10):
                col = col + 1
                plan = ws4.cell(row=i, column=j)
                aba.cell(row=lin + 4, column=col).value = plan.value
                if aba.cell(row=lin + 4, column=col).fill.start_color.index == '00000000':
                    cor = ws4.cell(row=i, column=j).fill.start_color.index
                    if cor != '00000000':
                        aba.cell(row=lin + 4, column=col).fill = PatternFill(fgColor=cor, fill_type='solid')

        col = 0
        for j in range(1, 9):
            col = col + 1
            aba.cell(row=lin + 5, column=col).fill = PatternFill(fgColor="808080", fill_type="solid")
    except Exception:
        pass

    try:
        for i in range(com[4], fim[4]):
            lin = lin + 1
            col = 0
            for j in range(1, 10):
                col = col + 1
                plan = ws5.cell(row=i, column=j)
                aba.cell(row=lin + 5, column=col).value = plan.value
                if aba.cell(row=lin + 5, column=col).fill.start_color.index == '00000000':
                    cor = ws5.cell(row=i, column=j).fill.start_color.index
                    if cor != '00000000':
                        aba.cell(row=lin + 5, column=col).fill = PatternFill(fgColor=cor, fill_type='solid')

        col = 0
        for j in range(1, 9):
            col = col + 1
            aba.cell(row=lin + 6, column=col).fill = PatternFill(fgColor="808080", fill_type="solid")
    except Exception:
        pass

    try:
        for i in range(com[5], fim[5]):
            lin = lin + 1
            col = 0
            for j in range(1, 10):
                col = col + 1
                plan = ws6.cell(row=i, column=j)
                aba.cell(row=lin + 6, column=col).value = plan.value
                if aba.cell(row=lin + 6, column=col).fill.start_color.index == '00000000':
                    cor = ws6.cell(row=i, column=j).fill.start_color.index
                    if cor != '00000000':
                        aba.cell(row=lin + 6, column=col).fill = PatternFill(fgColor=cor, fill_type='solid')

        col = 0
        for j in range(1, 9):
            col = col + 1
            aba.cell(row=lin + 7, column=col).fill = PatternFill(fgColor="808080", fill_type="solid")
    except Exception:
        pass


def salvar():
    openpyxl.Workbook.save(planilha, filename=roteiro)


# abrir planilha e remover abas padrao excell vazias
planilha = load_workbook(filename='C://Users/user/Local/NomeFinaldoArquivo')
try:
    aba1 = planilha['Plan2']
    planilha.remove(aba1)
except Exception:
    print()

try:
    aba2 = planilha['Plan3']
    planilha.remove(aba2)
except Exception:
    print()

# declarar aba ativa
aba = planilha['Plan1']

# pegar dados dos carros para começar formatação
for i in range(0, 300):
    nNota = aba.cell(row=i + 1, column=1).value
    placa = aba.cell(row=i + 1, column=2).value
    if nNota == 1:
        carro.append(placa)
        com.append(i + 1)
    if nNota is None:
        break

for i in range(0, 300):
    nNota = aba.cell(row=i + 1, column=1).value
    if nNota == 1 and i + 1 != 2:
        fim.append(i + 1)
    elif nNota is None:
        fim.append(i + 1)
        break

# formatação da planilha / remoção dos dados inuteis
for i in range(0, 300):
    descr = aba.cell(row=i + 1, column=13).value
    if descr is None:
        break
    else:
        aba.cell(row=i + 1, column=13).value = None

worksheet.Worksheet.delete_cols(aba, 3)
worksheet.Worksheet.delete_cols(aba, 4)
worksheet.Worksheet.delete_cols(aba, 4)
worksheet.Worksheet.delete_cols(aba, 7)
worksheet.Worksheet.delete_cols(aba, 7)

# formatação da planilha / separando carros
ws1 = planilha.create_sheet('SheetA')
try:
    ws1.title = carro[0]
    lin = 0
    col = 0
    for i in range(com[0], fim[0]):
        lin = lin + 1
        col = 0
        for j in range(1, 10):
            col = col + 1
            planP = aba.cell(row=i, column=j)
            ws1.cell(row=lin, column=col).value = planP.value
except Exception:
    planilha.remove(ws1)

ws2 = planilha.create_sheet('SheetA')
try:
    ws2.title = carro[1]
    lin = 0
    col = 0
    for i in range(com[1], fim[1]):
        lin = lin + 1
        col = 0
        for j in range(1, 10):
            col = col + 1
            planP = aba.cell(row=i, column=j)
            ws2.cell(row=lin, column=col).value = planP.value

except Exception:
    planilha.remove(ws2)

ws3 = planilha.create_sheet('SheetA')
try:
    ws3.title = carro[2]
    lin = 0
    col = 0
    for i in range(com[2], fim[2]):
        lin = lin + 1
        col = 0
        for j in range(1, 10):
            col = col + 1
            planP = aba.cell(row=i, column=j)
            ws3.cell(row=lin, column=col).value = planP.value
except Exception:
    planilha.remove(ws3)

ws4 = planilha.create_sheet('SheetA')
try:
    ws4.title = carro[3]
    lin = 0
    col = 0
    for i in range(com[3], fim[3]):
        lin = lin + 1
        col = 0
        for j in range(1, 10):
            col = col + 1
            planP = aba.cell(row=i, column=j)
            ws4.cell(row=lin, column=col).value = planP.value
except Exception:
    planilha.remove(ws4)

ws5 = planilha.create_sheet('SheetA')
try:
    ws5.title = carro[4]
    lin = 0
    col = 0
    for i in range(com[4], fim[4]):
        lin = lin + 1
        col = 0
        for j in range(1, 10):
            col = col + 1
            planP = aba.cell(row=i, column=j)
            ws5.cell(row=lin, column=col).value = planP.value
except Exception:
    planilha.remove(ws5)

ws6 = planilha.create_sheet('SheetA')
try:
    ws6.title = carro[5]
    lin = 0
    col = 0
    for i in range(com[5], fim[5]):
        lin = lin + 1
        col = 0
        for j in range(1, 10):
            col = col + 1
            planP = aba.cell(row=i, column=j)
            ws6.cell(row=lin, column=col).value = planP.value
except Exception:
    planilha.remove(ws6)

# pegando novos com e fim para formatação da planilha final
com.clear()
fim.clear()

for i in range(0, 300):
    nNota = ws1.cell(row=i + 1, column=1).value
    if nNota == 1:
        com.append(i + 1)

for i in range(0, 300):
    nNota = ws1.cell(row=i + 1, column=1).value
    if nNota is None:
        fim.append(i + 1)
        break

for i in range(0, 300):
    nNota = ws2.cell(row=i + 1, column=1).value
    if nNota == 1:
        com.append(i + 1)
    if nNota is None:
        break

for i in range(0, 300):
    nNota = ws2.cell(row=i + 1, column=1).value
    if nNota is None:
        fim.append(i + 1)
        break

for i in range(0, 300):
    nNota = ws3.cell(row=i + 1, column=1).value
    if nNota == 1:
        com.append(i + 1)
    if nNota is None:
        break

for i in range(0, 300):
    nNota = ws3.cell(row=i + 1, column=1).value
    if nNota is None:
        fim.append(i + 1)
        break

for i in range(0, 300):
    nNota = ws4.cell(row=i + 1, column=1).value
    if nNota == 1:
        com.append(i + 1)
    if nNota is None:
        break

for i in range(0, 300):
    nNota = ws4.cell(row=i + 1, column=1).value
    if nNota is None:
        fim.append(i + 1)
        break

for i in range(0, 300):
    nNota = ws5.cell(row=i + 1, column=1).value
    if nNota == 1:
        com.append(i + 1)
    if nNota is None:
        break

for i in range(0, 300):
    nNota = ws5.cell(row=i + 1, column=1).value
    if nNota is None:
        fim.append(i + 1)
        break

for i in range(0, 300):
    nNota = ws6.cell(row=i + 1, column=1).value
    if nNota == 1:
        com.append(i + 1)
    if nNota is None:
        break

for i in range(0, 300):
    nNota = ws6.cell(row=i + 1, column=1).value
    if nNota is None:
        fim.append(i + 1)
        break

# Salvando numero das notas do carro
try:
    for i in range(com[0], fim[0]):
        str(notas1.append(ws1.cell(row=i, column=3).value))


except Exception:
    pass

try:
    for i in range(com[1], fim[1]):
        notas2.append(ws2.cell(row=i, column=3).value)


except Exception:
    pass

try:
    for i in range(com[2], fim[2]):
        notas3.append(ws3.cell(row=i, column=3).value)


except Exception:
    pass

try:
    for i in range(com[3], fim[3]):
        notas4.append(ws4.cell(row=i, column=3).value)


except Exception:
    pass

try:
    for i in range(com[4], fim[4]):
        notas5.append(ws5.cell(row=i, column=3).value)


except Exception:
    pass

try:
    for i in range(com[5], fim[5]):
        notas6.append(ws6.cell(row=i, column=3).value)


except Exception:
    pass

check()
formatPlan()
salvar()
