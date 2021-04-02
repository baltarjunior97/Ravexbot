from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
from datetime import date
import utils
import openpyxl

# Variaveis
carro = []
com = []
fim = []
notas = [[]]
feitas = []
hoje = date.today().strftime("%d.%m.%Y")
roteiro = str("C://Users/user/Local/NomeFinaldoArquivo")

# abrir planilha e remover abas padrao excell vazias
planilha = load_workbook(filename="C://Users/user/Local/NomeFinaldoArquivo")
try:
    aba1 = planilha['Plan2']
    planilha.remove(aba1)
except Exception:
    pass

try:
    aba2 = planilha['Plan3']
    planilha.remove(aba2)
except Exception:
    pass

# declarar aba ativa
aba = planilha['Plan1']

# pegar dados dos carros para começar formatação
for i in range(0, 300):
    nNota = aba.cell(row=i + 1, column=1).value
    placa = aba.cell(row=i + 1, column=2).value
    if nNota == 1:
        carro.append(placa)
        com.append(i + 1)
    if nNota is None:
        break

for i in range(0, 300):
    nNota = aba.cell(row=i + 1, column=1).value
    if nNota == 1 and i + 1 != 2:
        fim.append(i + 1)
    elif nNota is None:
        fim.append(i + 1)
        break

# formatação da planilha / remoção dos dados inuteis
for i in range(0, 300):
    descr = aba.cell(row=i + 1, column=13).value
    if descr is None:
        break
    else:
        aba.cell(row=i + 1, column=13).value = None

worksheet.Worksheet.delete_cols(aba, 3)
worksheet.Worksheet.delete_cols(aba, 4)
worksheet.Worksheet.delete_cols(aba, 4)
worksheet.Worksheet.delete_cols(aba, 7)
worksheet.Worksheet.delete_cols(aba, 7)

ws = ['ws1', 'ws2', 'ws3', 'ws4', 'ws5', 'ws6']
for i in range(5):
    ws[i] = planilha.create_sheet(f'Sheet{i+1}')
    try:
        ws[i].title = carro[i]
        lin = 0
        col = 0
        for y in range(com[i], fim[i]):
            lin = lin+1
            col = 0
            for j in range (1, 10):
                col = col+1
                planP = aba.cell(row=y, column=j)
                ws[i].cell(row=lin, column=col).value = planP.value
    except Exception:
        planilha.remove(ws[i])

com.clear()
fim.clear()

for i in range(5):
    for y in range(0, 300):
        nNota = ws[i].cell(row=y+1, column=1).value
        if nNota == 1:
            com.append(y+1)
        if nNota is None:
            fim.append(y+1)
            break
for i in range(5):
    try:
        for y in range(com[i], fim[i]):
            str(notas[i].append(ws[i].cell(row=y, column=3).value))
    except Exception:
        continue

utils.formatPlan(com, fim, ws, aba)
openpyxl.Workbook.save(planilha, filename=roteiro)

#utils.semCarimbo(carro, notas, ws, com, fim)
#utils.reentrega(carro, notas, ws, com, fim)
#utils.devolucao(carro, notas, ws, com, fim)
#utils.anotacao(carro, notas, ws, com, fim)
#utils.check(carro, com, fim, ws, feitas)
utils.salvar(planilha, roteiro)
