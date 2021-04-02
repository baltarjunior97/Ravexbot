from openpyxl.styles import PatternFill
import openpyxl


def formatPlan(com, fim, ws, aba):
    lin =0
    for i in range(5):
        try:
            for y in range(com[i], fim[i]):
                lin = lin +1
                col = 0
                for j in range(1, 10):
                    col = col +1
                    plan = ws[i].cell(row=y, column=j)
                    aba.cell(row=lin+1+i, column=col).value = plan.value
                    if aba.cell(row=lin+1+i, column=col).fill.start_color.index == '00000000':
                        cor = ws[i].cell(row=y, column=j).fill.start_color.index
                        if cor != '00000000':
                            aba.cell(row=lin+1+i, column=col).fill = PatternFill(fgColor=cor, fill_type='solid')
            col = 0
            for j in range(1, 9):
                col = col+1
                aba.cell(row=lin+2+i, column=col).fill = PatternFill(fgColor="808080", fill_type="solid")
                aba.cell(row=lin+2+i, column=col).value = None
        except Exception as e:
            continue


def salvar(planilha, roteiro):
    openpyxl.Workbook.save(planilha, filename=roteiro)


def semCarimbo(carro, notas, ws, com, fim):
    nNota = 0
    car = str(input('Qual carro? '))
    while True:
        if car not in carro:
            print('Carros Disponiveis:')
            for y in range(0, len(carro)):
                print(carro[y])
            car = str(input('Qual carro? '))
        elif car in carro:
            break
    xpos = carro.index(car)
    nNota = int(input(f'Qual nota do carro {car}? '))
    while True:
        if nNota not in notas[xpos]:
            nNota = int(input(f'Qual nota do carro {car}? '))
        elif nNota in notas[xpos]:
            break
    note = 'sem carimbo'
    for i in range(com[xpos], fim[xpos]):
        if ws[xpos].cell(row=i, column=3).value == nNota:
            ws[xpos].cell(row=i, column=8).value = note


def reentrega(carro, notas, ws, com, fim):
    nNota = 0
    car = str(input('Qual carro teve a reentrega? '))
    while True:
        if car not in carro:
            print('Carros Disponiveis:')
            for y in range(0, len(carro)):
                print(carro[y])
            car = str(input('Qual carro teve a reentrega? '))
        elif car in carro:
            break
    xpos = carro.index(car)
    while True:
        if nNota not in notas[xpos]:
            nNota = int(input(f'Qual nota do carro {car}? '))
        elif nNota in notas[xpos]:
            break
    note = str(input('Digite o motivo: '))

    for y in range(com[xpos], fim[xpos]):
        if ws[xpos].cell(row=y, column=3).value == nNota:
            ws[xpos].cell(row=y, column=8).value = note
            for j in range(1, 9):
                ws[xpos].cell(row=y, column=j).fill = PatternFill(fgColor="00FFFF00", fill_type = "solid")


def devolucao(carro, notas, ws, com, fim):
    nNota = 0
    dev = 0

    dev = str(input('Devolução parcial(1) ou devolução total(2)? '))
    while True:
        if dev == '1' or dev == 'devolução parcial' or dev == 'devolucao parcial':
            dev = 'devolução parcial'
            break
        elif dev == '2' or dev == 'devolução total' or dev == 'devoucao total':
            dev = 'devolução total'
            break
        else:
            dev = str(input('Devolução parcial(1) ou devolução total(2)? '))

    car = str(input(f'Qual carro teve a {dev}? '))
    while True:
        if car not in carro:
            print('Carros Disponiveis:')
            for y in range(0, len(carro)):
                print(carro[y])
            car = str(input(f'Qual carro teve a {dev}? '))
        elif car in carro:
            break
    xpos = carro.index(car)
    while True:
        if nNota not in notas[xpos]:
            nNota = int(input(f'Qual nota do carro {car}? '))
        elif nNota in notas[xpos]:
            break
    note = str(input('Digite o motivo: '))
    for y in range(com[xpos], fim[xpos]):
        if ws[xpos].cell(row=y, column=3).value == nNota:
            ws[xpos].cell(row=y, column=8).value = note
            for j in range(1, 9):
                ws[xpos].cell(row=y, column=j).fill = PatternFill(fgColor="00FF0000", fill_type = "solid")


def anotacao(carro, notas, ws, com, fim):
    nNota = 0
    car = str(input('Qual carro vc quer fazer a anotaçao? '))
    while True:
        if car not in carro:
            print('Carros Disponiveis:')
            for y in range(0, len(carro)):
                print(carro[y])
            car = str(input('Qual carro vc quer fazer a anotaçao? '))
        elif car in carro:
            break
    xpos = carro.index(car)
    while True:
        if nNota not in notas[xpos]:
            nNota = int(input(f'Qual nota do carro {car}? '))
        elif nNota in notas[xpos]:
            break
    note = str(input('Digite a anotação: '))
    for y in range(com[xpos], fim[xpos]):
        if ws[xpos].cell(row=y, column=3).value == nNota:
            ws[xpos].cell(row=y, column=8).value = note


def check(carro, com, fim, ws, feitas):
    for i in range(0, len(carro)):
        xpos = i
        while True:
            try:
                x = int(input(f'até que nota o {carro[i]} fez? '))
                if x >= com[i] and x <= fim[i]:
                    feitas.append(x)
                    break
            except Exception as e:
                print(e)

        for y in range(com[xpos], fim[xpos]):
            for j in range(1, 9):
                if ws[xpos].cell(row=y, column=j).fill.start_color.index == '00000000':
                    ws[xpos].cell(row=y, column=j).fill = PatternFill(fgColor="0000FF00", fill_type = "solid")