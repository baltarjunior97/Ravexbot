from openpyxl import cell, load_workbook
import openpyxl
from openpyxl.worksheet import worksheet
from datetime import date
from openpyxl.styles import PatternFill
import schedule


def realizada():
    check = str(input("Gostaria de ajustar o monitoramento?(s/n)")).lower()
    if check == 's':
        for i in range(0, len(carro)):
            x = int(input(f'Em qual nota o carro {carro[i]} está? '))
            if x > fim[i] or x < com[i]:
                print(f'carro {carro[i]} só vai até a nota {fim[i]}')
                x = int(input(f'Em qual nota o carro {carro[i]} está? '))
            else: 
                feitas.append(x+1)
        for i in range(0, len(carro)):
            for e in range(com[i], fim[i]):
                for rows in sheet.iter_rows(com[i], fim[i]):
                    for cell in rows:
                        cell.fill = PatternFill(bgColor="1EE030", fill_type = "solid")
    if check == 'n':
     print('OK')
    else: 
     check = str(input('Opção invalida, Digite novamente: '))

feitas = []
carro = []
com = []
fim = []
carro.clear()
com.clear()
fim.clear()
feitas.clear()

hoje = date.today()

d1 = hoje.strftime("%d.%m.%Y")

wb = load_workbook('C:\\Users\djalm\Downloads\teste.xlsx')

sheet = wb['Plan1'] 

for i in range(0, 300):
    x = sheet.cell(row=i+1, column=1).value
    y = sheet.cell(row=i+1, column=2).value
    if x == 1 :
        carro.append(y)
        com.append(i+1)

for i in range(0, 300):
    x = sheet.cell(row=i+1, column=1).value
    y = sheet.cell(row=i+1, column=2).value
    if x == 1 and i+1 != 2:
        fim.append(i)
    elif x == None:
        fim.append(i)
        break

for i in range(0, 300):
    x = sheet.cell(row=i+1, column=13).value
    if x == None:
        break
    else:
        sheet.cell(row=i+1, column=13).value = None

worksheet.Worksheet.delete_cols(sheet, 3)
worksheet.Worksheet.delete_cols(sheet, 4)
worksheet.Worksheet.delete_cols(sheet, 4)
worksheet.Worksheet.delete_cols(sheet, 7)
worksheet.Worksheet.delete_cols(sheet, 7)

realizada()
schedule.every().day.at('10:50').do(realizada)
        
schedule.every().day.at('13:50').do(realizada)

schedule.every().day.at('16:50').do(realizada)

roteiro =str("C://Users/djalm/Downloads/roteiro{}{}".format(d1, ".xlsx"))    
openpyxl.Workbook.save(wb, filename=roteiro)