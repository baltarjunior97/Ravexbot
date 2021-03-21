from openpyxl import load_workbook
import openpyxl
from openpyxl.worksheet import worksheet
from datetime import date
from openpyxl.styles import PatternFill
import schedule

feitas = []
carro = []
com = []
fim = []
carro.clear()
com.clear()
fim.clear()
feitas.clear()
wsl1 = 1
col = 0
lin = 0
hoje = date.today()

d1 = hoje.strftime("%d.%m.%Y")

wb = load_workbook('C:\\Users\djalm\Downloads\teste.xlsx')
try:
    sheet1 = wb['Plan2']
    wb.remove_sheet(sheet1)
except:
    print()

try:
    sheet2 = wb['Plan3']
    wb.remove_sheet(sheet2)
except:
    print()

sheet = wb['Plan1'] 

for i in range(0, 300):
    x = sheet.cell(row=i+1, column=1).value
    y = sheet.cell(row=i+1, column=2).value
    if x == 1 :
        carro.append(y)
        com.append(i+1)

for i in range(0, 300):
    x = sheet.cell(row=i+1, column=1).value
    y = sheet.cell(row=i+1, column=2).value
    if x == 1 and i+1 != 2:
        fim.append(i+1)
    elif x == None:
        fim.append(i+1)
        break

for i in range(0, 300):
    x = sheet.cell(row=i+1, column=13).value
    if x == None:
        break
    else:
        sheet.cell(row=i+1, column=13).value = None

worksheet.Worksheet.delete_cols(sheet, 3)
worksheet.Worksheet.delete_cols(sheet, 4)
worksheet.Worksheet.delete_cols(sheet, 4)
worksheet.Worksheet.delete_cols(sheet, 7)
worksheet.Worksheet.delete_cols(sheet, 7)
   
ws1 = wb.create_sheet('SheetA')
try:
    ws1.title = carro[0]
    lin = 0
    col = 0
    for i in range(com[0], fim[0]):
        lin = lin+1
        col = 0
        for j in range(1 ,10):
            col = col+1
            c = sheet.cell(row=i, column=j)
            ws1.cell(row=lin, column=col).value = c.value
            
except:
    wb.remove_sheet(ws1)

ws2 = wb.create_sheet('SheetA')
try:   
    ws2.title = carro[1]
    lin = 0
    col = 0
    for i in range(com[1], fim[1]):
        lin = lin+1
        col = 0
        for j in range(1 ,10):
            col = col+1
            c = sheet.cell(row=i, column=j)
            ws2.cell(row=lin, column=col).value = c.value
        
except:
    wb.remove_sheet(ws2)

ws3 = wb.create_sheet('SheetA')
try:    
    ws3.title = carro[2]
    lin = 0
    col = 0
    for i in range(com[2], fim[2]):
        lin = lin+1
        col = 0
        for j in range(1 ,10):
            col = col+1
            c = sheet.cell(row=i, column=j)
            ws3.cell(row=lin, column=col).value = c.value
except:
    wb.remove_sheet(ws3)

ws4 = wb.create_sheet('SheetA')
try:    
    ws4.title = carro[3]
    lin = 0
    col = 0
    for i in range(com[3], fim[3]):
        lin = lin+1
        col = 0
        for j in range(1 ,10):
            col = col+1
            c = sheet.cell(row=i, column=j)
            ws4.cell(row=lin, column=col).value = c.value
except:
    wb.remove_sheet(ws4)

ws5 = wb.create_sheet('SheetA')
try:    
    ws5.title = carro[4]
    lin = 0
    col = 0
    for i in range(com[4], fim[4]):
        lin = lin+1
        col =0
        for j in range(1 ,10):
            col = col+1
            c = sheet.cell(row=i, column=j)
            ws5.cell(row=lin, column=col).value = c.value
except:
    wb.remove_sheet(ws5)

ws6 = wb.create_sheet('SheetA')
try:    
    ws6.title = carro[5]
    lin = 0
    col = 0
    for i in range(com[5], fim[5]):
        lin = lin+1
        col = 0
        for j in range(1 ,10):
            col = col+1
            c = sheet.cell(row=i, column=j)
            ws6.cell(row=lin, column=col).value = c.value
except:
    wb.remove_sheet(ws6)

com.clear()
fim.clear()

for i in range(0, 300):
    x = ws1.cell(row=i+1, column=1).value
    if x == 1 :
        com.append(i+1)

for i in range(0, 300):
    x = ws1.cell(row=i+1, column=1).value
    if x == None:
        fim.append(i+1)
        break

for i in range(0, 300):
    x = ws2.cell(row=i+1, column=1).value
    y = ws2.cell(row=i+1, column=2).value
    if x == 1 :
        com.append(i+1)

for i in range(0, 300):
    x = ws2.cell(row=i+1, column=1).value
    if x == None:
        fim.append(i+1)
        break

for i in range(0, 300):
    x = ws3.cell(row=i+1, column=1).value
    y = ws3.cell(row=i+1, column=2).value
    if x == 1 :
        com.append(i+1)

for i in range(0, 300):
    x = ws3.cell(row=i+1, column=1).value
    if x == None:
        fim.append(i+1)
        break

for i in range(0, 300):
    x = ws4.cell(row=i+1, column=1).value
    y = ws4.cell(row=i+1, column=2).value
    if x == 1 :
        com.append(i+1)

for i in range(0, 300):
    x = ws4.cell(row=i+1, column=1).value
    if x == None:
        fim.append(i+1)
        break

for i in range(0, 300):
    x = ws5.cell(row=i+1, column=1).value
    y = ws5.cell(row=i+1, column=2).value
    if x == 1 :
        com.append(i+1)

for i in range(0, 300):
    x = ws5.cell(row=i+1, column=1).value
    if x == None:
        fim.append(i+1)
        break

for i in range(0, 300):
    x = ws6.cell(row=i+1, column=1).value
    y = ws6.cell(row=i+1, column=2).value
    if x == 1 :
        com.append(i+1)

for i in range(0, 300):
    x = ws6.cell(row=i+1, column=1).value
    if x == None:
        fim.append(i+1)
        break

teste = wb.create_sheet('SheetA')
lin = 0
col = 0

for j in range(1,10):
    col = col+1
    teste.cell(row=lin+1, column=col).fill = PatternFill(bgColor="111111", fill_type = "solid")
for i in range(com[0], fim[0]):
    lin = lin+1
    col = 0
    for j in range(1, 10):
        col = col+1
        c = ws1.cell(row=i, column=j)
        teste.cell(row=lin+1, column=col).value = c.value
col = 0
for j in range(1,10):
    col = col+1
    teste.cell(row=lin+2, column=col).fill = PatternFill(bgColor="111111", fill_type = "solid")

try:
    for i in range(com[1], fim[1]):
        lin = lin+1
        col = 0
        for j in range(1, 10):
            col = col+1
            c = ws2.cell(row=i, column=j)
            teste.cell(row=lin+2, column=col).value = c.value

    col = 0
    for j in range(1,10):
        col = col+1
        teste.cell(row=lin+3, column=col).fill = PatternFill(bgColor="A1A1A1", fill_type = "solid")
except:
    print()

try:
    for i in range(com[2], fim[2]):
        lin = lin+1
        col = 0
        for j in range(1, 10):
            col = col+1
            c = ws3.cell(row=i, column=j)
            teste.cell(row=lin+3, column=col).value = c.value

    col = 0
    for j in range(1,10):
        col = col+1
        teste.cell(row=lin+4, column=col).fill = PatternFill(bgColor="A1A1A1", fill_type = "solid")
except:
            print()
try:
    for i in range(com[3], fim[3]):
        lin = lin+1
        col = 0
        for j in range(1, 10):
            col = col+1
            c = ws4.cell(row=i, column=j)
            teste.cell(row=lin+4, column=col).value = c.value

    col = 0
    for j in range(1,10):
        col = col+1
        teste.cell(row=lin+5, column=col).fill = PatternFill(bgColor="A1A1A1", fill_type = "solid")
except:
    print()

try:
    for i in range(com[4], fim[4]):
        lin = lin+1
        col = 0
        for j in range(1, 10):
            col = col+1
            c = ws5.cell(row=i, column=j)
            teste.cell(row=lin+5, column=col).value = c.value

    col = 0
    for j in range(1,10):
        col = col+1
        teste.cell(row=lin+6, column=col).fill = PatternFill(bgColor="A1A1A1", fill_type = "solid")
except:
    print()

try:
    for i in range(com[5], fim[5]):
        lin = lin+1
        col = 0
        for j in range(1, 10):
            col = col+1
            c = ws6.cell(row=i, column=j)
            teste.cell(row=lin+6, column=col).value = c.value

    col = 0
    for j in range(1,10):
        col = col+1
        teste.cell(row=lin+7, column=col).fill = PatternFill(bgColor="A1A1A1", fill_type = "solid")
except:
    print()
'''
schedule.every().day.at('10:50').do(realizada)
        
schedule.every().day.at('13:50').do(realizada)

schedule.every().day.at('16:50').do(realizada)
'''
roteiro =str("C://Users/djalm/Downloads/roteiro{}{}".format(d1, ".xlsx"))    
openpyxl.Workbook.save(wb, filename=roteiro)